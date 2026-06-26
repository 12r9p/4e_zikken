import os
import math
import warnings
import openpyxl
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm

# Ignore warnings (including font-related ones)
warnings.filterwarnings('ignore')

# Set clean aesthetic style
plt.rcParams['figure.facecolor'] = 'white'
plt.rcParams['axes.facecolor'] = '#FCFCFC'
plt.rcParams['grid.color'] = '#EBEBEB'
plt.rcParams['grid.linestyle'] = '--'
plt.rcParams['grid.linewidth'] = 0.8
plt.rcParams['font.size'] = 11
plt.rcParams['axes.edgecolor'] = '#CCCCCC'
plt.rcParams['axes.linewidth'] = 1.0
plt.rcParams['xtick.color'] = '#444444'
plt.rcParams['ytick.color'] = '#444444'
plt.rcParams['axes.labelcolor'] = '#222222'
plt.rcParams['axes.titlecolor'] = '#111111'

# Setup directory paths relative to this script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR) # /Users/takumi/Documents/4e_zikken/6_kato
DATA_DIR = os.path.join(BASE_DIR, 'data')
ASSETS_DIR = os.path.join(BASE_DIR, 'assets')
XLSX_PATH = os.path.join(BASE_DIR, '加藤実験.xlsx')

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(ASSETS_DIR, exist_ok=True)

# Select a Japanese font if available
jp_font_name = 'sans-serif'
for f in fm.fontManager.ttflist:
    if any(name in f.name for name in ['Hiragino', 'IPA', 'YuGothic', 'MS Gothic', 'AppleGothic', 'Takao', 'Osaka']):
        jp_font_name = f.name
        break
plt.rcParams['font.family'] = jp_font_name

print(f"Using font: {jp_font_name}")

# Load Excel workbook
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

# ----------------------------------------------------
# 1. Load Characteristic (実負荷試験)
# ----------------------------------------------------
print("Processing load characteristic...")
ws_load = wb['同期発電機の実負荷試験']
load_data = []

# Row 3 to 14
# Column G (index 6): Terminal Voltage V_L [V]
# Column H (index 7): Load Current I_L [A]
for r in range(3, 15):
    v_l = ws_load.cell(row=r, column=7).value
    i_l = ws_load.cell(row=r, column=8).value
    if v_l is not None and i_l is not None:
        load_data.append({'Load Current I_L [A]': float(i_l), 'Terminal Voltage V_L [V]': float(v_l)})

df_load = pd.DataFrame(load_data)
# Sort by Load Current ascending
df_load = df_load.sort_values(by='Load Current I_L [A]').reset_index(drop=True)
df_load.to_csv(os.path.join(DATA_DIR, 'load_characteristic.csv'), index=False)

# Plot Load Characteristic
fig, ax = plt.subplots(figsize=(6.5, 4.5))
ax.plot(df_load['Load Current I_L [A]'], df_load['Terminal Voltage V_L [V]'], 
        marker='o', color='#2563EB', linewidth=2, markersize=6, label='Measured $V_L$')
# Mark rated voltage 200V and no-load voltage 210.5V
ax.axhline(200.0, color='#94A3B8', linestyle=':', linewidth=1.5, label='Rated Voltage 200V')
ax.scatter([0.0], [210.5], color='#EF4444', s=60, zorder=5, label='No-load Voltage $V_0$ = 210.5V')
ax.scatter([5.75], [197.5], color='#10B981', s=60, zorder=5, label='Max Load Point')

ax.set_title('Synchronous Generator Load Characteristic\n(同期発電機の実負荷特性)', fontsize=12, pad=10)
ax.set_xlabel('Load Current $I_L$ [A]\n(負荷電流)', fontsize=10)
ax.set_ylabel('Terminal Voltage $V_L$ [V]\n(端子電圧)', fontsize=10)
ax.set_xlim(-0.5, 6.5)
ax.set_ylim(190, 220)
ax.grid(True)
ax.legend(loc='best', frameon=True, facecolor='white', edgecolor='#E2E8F0')

# Annotation for voltage regulation
eps_text = r"$\epsilon = \frac{V_0 - V_n}{V_n} \times 100 = \frac{210.5 - 200}{200} \times 100 = 5.25\%$"
ax.text(1.0, 193.0, eps_text, fontsize=10.5, bbox=dict(facecolor='white', alpha=0.8, edgecolor='#CBD5E1'))

plt.tight_layout()
fig.savefig(os.path.join(ASSETS_DIR, 'load_characteristic.png'), dpi=300)
plt.close(fig)


# ----------------------------------------------------
# 2. No-load Characteristic (無負荷飽和試験)
# ----------------------------------------------------
print("Processing no-load characteristic...")
ws_noload = wb['同期発電機の無負荷試験']

# Increasing: rows 4 to 25
inc_data = []
for r in range(4, 26):
    i_f = ws_noload.cell(row=r, column=4).value
    v_0l = ws_noload.cell(row=r, column=5).value
    if i_f is not None and v_0l is not None:
        inc_data.append({'Direction': 'Increasing', 'Field Current I_F [A]': float(i_f), 'Terminal Voltage V_0L [V]': float(v_0l)})

# Decreasing: rows 26 to 48
dec_data = []
for r in range(26, 49):
    i_f = ws_noload.cell(row=r, column=4).value
    v_0l = ws_noload.cell(row=r, column=5).value
    if i_f is not None and v_0l is not None:
        dec_data.append({'Direction': 'Decreasing', 'Field Current I_F [A]': float(i_f), 'Terminal Voltage V_0L [V]': float(v_0l)})

df_inc = pd.DataFrame(inc_data)
df_dec = pd.DataFrame(dec_data)
df_noload = pd.concat([df_inc, df_dec]).reset_index(drop=True)
df_noload.to_csv(os.path.join(DATA_DIR, 'no_load_characteristic.csv'), index=False)

# Plot No-load Characteristic
fig, ax = plt.subplots(figsize=(6.5, 4.5))
ax.plot(df_inc['Field Current I_F [A]'], df_inc['Terminal Voltage V_0L [V]'], 
        marker='o', color='#0D9488', linewidth=2, markersize=5, label='Increasing Branch (往路)')
ax.plot(df_dec['Field Current I_F [A]'], df_dec['Terminal Voltage V_0L [V]'], 
        marker='x', color='#EA580C', linewidth=2, markersize=5, label='Decreasing Branch (復路)')

ax.set_title('Open-Circuit Saturation Curve\n(無負荷飽和特性曲線)', fontsize=12, pad=10)
ax.set_xlabel('Field Current $I_F$ [A]\n(界磁電流)', fontsize=10)
ax.set_ylabel('Terminal Voltage $V_{0L}$ [V]\n(無負荷端子電圧)', fontsize=10)
ax.grid(True)
ax.legend(loc='lower right', frameon=True, facecolor='white', edgecolor='#E2E8F0')
plt.tight_layout()
fig.savefig(os.path.join(ASSETS_DIR, 'no_load_characteristic.png'), dpi=300)
plt.close(fig)


# ----------------------------------------------------
# 3. Short-circuit Characteristic (三相短絡試験)
# ----------------------------------------------------
print("Processing short-circuit characteristic...")
sc_data = []

# Rows 4 to 16
# Column H (index 7): Field Current I_F [A]
# Column I, J, K (indices 8, 9, 10): Short-circuit Currents
# Column L (index 11): Average Short-circuit Current
for r in range(4, 17):
    i_f = ws_noload.cell(row=r, column=8).value
    i_1 = ws_noload.cell(row=r, column=9).value
    i_2 = ws_noload.cell(row=r, column=10).value
    i_3 = ws_noload.cell(row=r, column=11).value
    if i_f is not None and all(v is not None for v in [i_1, i_2, i_3]):
        avg_i = (float(i_1) + float(i_2) + float(i_3)) / 3.0
        sc_data.append({
            'Field Current I_F [A]': float(i_f),
            'Short Circuit Current I_L1 [A]': float(i_1),
            'Short Circuit Current I_L2 [A]': float(i_2),
            'Short Circuit Current I_L3 [A]': float(i_3),
            'Average Short Circuit Current I_s [A]': avg_i
        })

df_sc = pd.DataFrame(sc_data)
# Sort by Field Current ascending
df_sc = df_sc.sort_values(by='Field Current I_F [A]').reset_index(drop=True)
df_sc.to_csv(os.path.join(DATA_DIR, 'short_circuit_characteristic.csv'), index=False)

# Plot Short-circuit Characteristic
fig, ax = plt.subplots(figsize=(6.5, 4.5))
ax.scatter(df_sc['Field Current I_F [A]'], df_sc['Average Short Circuit Current I_s [A]'], 
           color='#4F46E5', marker='o', s=40, zorder=5, label='Measured Average $I_s$')

# Linear fit to show linear relationship
slope, intercept = np.polyfit(df_sc['Field Current I_F [A]'], df_sc['Average Short Circuit Current I_s [A]'], 1)
fit_x = np.linspace(0.0, df_sc['Field Current I_F [A]'].max() * 1.1, 100)
fit_y = slope * fit_x + intercept
ax.plot(fit_x, fit_y, color='#818CF8', linestyle='--', linewidth=1.5, label=f'Linear Fit: $I_s = {slope:.2f} I_F + {intercept:.2f}$')

ax.set_title('Three-Phase Short-Circuit Curve\n(三相短絡特性曲線)', fontsize=12, pad=10)
ax.set_xlabel('Field Current $I_F$ [A]\n(界磁電流)', fontsize=10)
ax.set_ylabel('Short-Circuit Current $I_s$ [A]\n(短絡電流)', fontsize=10)
ax.set_xlim(0, df_sc['Field Current I_F [A]'].max() * 1.15)
ax.set_ylim(0, df_sc['Average Short Circuit Current I_s [A]'].max() * 1.15)
ax.grid(True)
ax.legend(loc='upper left', frameon=True, facecolor='white', edgecolor='#E2E8F0')
plt.tight_layout()
fig.savefig(os.path.join(ASSETS_DIR, 'short_circuit_characteristic.png'), dpi=300)
plt.close(fig)


# ----------------------------------------------------
# 4. Synchronous Impedance Calculation (同期インピーダンス算出)
# ----------------------------------------------------
print("Processing synchronous impedance...")
# Compute for field current x = 0.5 to 3.0 step 0.1
z_data = []
for i in range(5, 31):
    x = i / 10.0
    # open_voltage = -2.0091*x^4 + 26.46385*x^3 - 127.996*x^2 + 295.27*x - 31.624
    v_open = -2.0091*(x**4) + 26.46385*(x**3) - 127.996*(x**2) + 295.27*x - 31.624
    # short_current = 8.45*x + 0.6227
    i_short = 8.45*x + 0.6227
    # Zs = v_open / (sqrt(3) * i_short)
    z_s = v_open / (math.sqrt(3.0) * i_short)
    z_data.append({
        'Field Current I_F [A]': x,
        'Open Circuit Voltage V_0 [V]': v_open,
        'Short Circuit Current I_s [A]': i_short,
        'Synchronous Impedance Z_s [Ohm]': z_s
    })

df_z = pd.DataFrame(z_data)
df_z.to_csv(os.path.join(DATA_DIR, 'synchronous_impedance.csv'), index=False)

# Plot Synchronous Impedance
fig, ax1 = plt.subplots(figsize=(7, 4.5))

color_z = '#7C3AED'
ax1.plot(df_z['Field Current I_F [A]'], df_z['Synchronous Impedance Z_s [Ohm]'], 
         marker='s', color=color_z, linewidth=2.5, markersize=5, label='Synchronous Impedance $Z_s$')
ax1.set_xlabel('Field Current $I_F$ [A]\n(界磁電流)', fontsize=10)
ax1.set_ylabel(r'Synchronous Impedance $Z_s$ [$\Omega$]' + '\n(同期インピーダンス)', color=color_z, fontsize=10)
ax1.tick_params(axis='y', labelcolor=color_z)
ax1.grid(True)

# Plot open-circuit voltage & short-circuit current on secondary axis to show why Zs drops
ax2 = ax1.twinx()
color_v = '#0D9488'
color_i = '#4F46E5'
ax2.plot(df_z['Field Current I_F [A]'], df_z['Open Circuit Voltage V_0 [V]'], 
         color=color_v, linestyle=':', linewidth=1.5, label='Formula-fitted $V_0$')
ax2.plot(df_z['Field Current I_F [A]'], df_z['Short Circuit Current I_s [A]'], 
         color=color_i, linestyle='-.', linewidth=1.5, label='Formula-fitted $I_s$')
ax2.set_ylabel('Voltage $V_0$ [V] / Current $I_s$ [A]', color='#333333', fontsize=10)
ax2.tick_params(axis='y', labelcolor='#333333')

# Combine legends
lines1, labels1 = ax1.get_legend_handles_labels()
lines2, labels2 = ax2.get_legend_handles_labels()
ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper right', frameon=True, facecolor='white', edgecolor='#E2E8F0')

ax1.set_title('Synchronous Impedance vs. Field Current\n(同期インピーダンスの界磁電流依存性)', fontsize=12, pad=10)
plt.tight_layout()
fig.savefig(os.path.join(ASSETS_DIR, 'synchronous_impedance.png'), dpi=300)
plt.close(fig)


# ----------------------------------------------------
# 5. Conceptual DC Motor Graph (直流電動機 速度-トルク特性 概念図)
# ----------------------------------------------------
print("Generating conceptual DC motor graph...")
# N = V / (kPhi) - (r / (k * K' * Phi^2)) * T
# Setup range of torque T
T = np.linspace(0, 45, 200)
V = 100.0
r = 1.0

# Flux values Phi
fluxes = [0.8, 1.0, 1.2]
labels = [
    r'Weak Flux $\Phi = 0.8$ (弱界磁)',
    r'Nominal Flux $\Phi = 1.0$ (標準界磁)',
    r'Strong Flux $\Phi = 1.2$ (強界磁)',
]
colors = ['#EF4444', '#10B981', '#3B82F6']

fig, ax = plt.subplots(figsize=(6.5, 4.5))

# Plot motor characteristics
for phi, label, color in zip(fluxes, labels, colors):
    # N_0 = V / phi, slope = r / phi^2
    N = V / phi - (r / (phi**2)) * T
    ax.plot(T, N, color=color, linewidth=2, label=label)

# Plot load torque curve: T_load = 0.0016 * N^2
# For plotting, N is vertical axis, T is horizontal axis, so we plot (0.0016 * N^2, N)
N_range = np.linspace(50, 130, 200)
T_load = 0.0016 * (N_range**2)
ax.plot(T_load, N_range, color='#6B7280', linestyle='--', linewidth=2, label='Load Torque $T_{load} = C N^2$ (負荷側)')

# Add operating points (intersections calculated in planning)
# Weak: N=100, T=16
ax.scatter([16.0], [100.0], color='#DC2626', s=50, zorder=5)
ax.annotate('A (weak flux)', xy=(16.0, 100.0), xytext=(18.0, 103.0),
            arrowprops=dict(arrowstyle="->", color='#DC2626', lw=0.8), fontsize=9, color='#DC2626')

# Nominal: N=87.7, T=12.3
ax.scatter([12.3], [87.7], color='#059669', s=50, zorder=5)
ax.annotate('B (nominal)', xy=(12.3, 87.7), xytext=(14.3, 90.7),
            arrowprops=dict(arrowstyle="->", color='#059669', lw=0.8), fontsize=9, color='#059669')

# Strong: N=76.5, T=9.8
ax.scatter([9.8], [76.5], color='#2563EB', s=50, zorder=5)
ax.annotate('C (strong flux)', xy=(9.8, 76.5), xytext=(11.8, 79.5),
            arrowprops=dict(arrowstyle="->", color='#2563EB', lw=0.8), fontsize=9, color='#2563EB')

ax.set_title('DC Shunt Motor Torque-Speed Characteristics\n(直流分巻電動機の速度-トルク特性概念図: 理論値)', fontsize=11, pad=10)
ax.set_xlabel('Torque $T$ [a.u.] (トルク)', fontsize=10)
ax.set_ylabel('Rotational Speed $N$ [a.u.] (回転速度)', fontsize=10)
ax.set_xlim(0, 35)
ax.set_ylim(40, 135)
ax.grid(True)
ax.legend(loc='upper right', frameon=True, facecolor='white', edgecolor='#E2E8F0')

# Warning: Conceptual label
ax.text(1.0, 45, "* Conceptual diagram for theoretical discussion.\n  No direct measurements found in Excel workbook.", 
        fontsize=8.5, color='#94A3B8', style='italic')

plt.tight_layout()
fig.savefig(os.path.join(ASSETS_DIR, 'dc_motor_speed_torque_concept.png'), dpi=300)
plt.close(fig)

print("All assets generated successfully!")
